import openpyxl as xl
wb=xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
cell=sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)

for row in range(1,sheet.max_row+1):
    print(row)
    